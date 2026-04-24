import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment


class ExcelGenerator:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = Workbook()

    def create_report(self, summary_data):
        ws1 = self.workbook.active
        ws1.title = 'Summary'
        self.add_sheet_data(ws1, summary_data)

        ws2 = self.workbook.create_sheet(title='Windows')
        self.add_sheet_data(ws2, summary_data.get('windows', []))

        ws3 = self.workbook.create_sheet(title='Doors')
        self.add_sheet_data(ws3, summary_data.get('doors', []))

        self.workbook.save(self.filename)

    def add_sheet_data(self, sheet, data):
        # Set headers
        headers = data[0].keys() if data else []
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(style='thin'))

        # Add data rows
        for row_num, row_data in enumerate(data, 2):
            for col_num, cell_value in enumerate(row_data.values(), 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        # Auto-fit columns
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter  # Get the column letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

# Example usage:
# generator = ExcelGenerator('report.xlsx')
# summary_data = {'windows': [{'Window Type': 'Casement', 'Count': 10}], 'doors': [{'Door Type': 'Interior', 'Count': 5}]}
# generator.create_report(summary_data)